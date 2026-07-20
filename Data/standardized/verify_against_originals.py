"""Independent verification: standardized tables vs ORIGINAL source files.
Reconstructs expected values directly from original/*.prism, original/*.xlsx and
the raw log zips (NOT via the build scripts) and asserts every mouse x day matches.
"""
import zipfile, json, re, csv, io, glob, os
from pathlib import Path
from datetime import date, datetime, timedelta
import openpyxl

BASE=Path(r"c:/Users/masba/OneDrive/Documents/Tumbly_paper/data")
STD=BASE/"standardized"
CYN, WIN = BASE/"Cynthia", BASE/"Winston"
TRF_START={"Cynthia":date(2026,6,25),"Winston":date(2026,6,20)}
TOL=1e-6
problems=[]
def check(name, cond, detail=""):
    print(("  PASS " if cond else "  FAIL ")+name+("" if cond else "  <-- "+detail))
    if not cond: problems.append(name+" :: "+detail)

def fnum(x):
    try: return float(x)
    except: return None

# ---------- load standardized ----------
# published data is anonymized (Lab A/Lab B); map back to source names so the
# reconstruction-from-originals below (which reads the Cynthia/Winston folders) matches.
_UNMAP = {"Lab A": "Cynthia", "Lab B": "Winston"}
def load_std(fn):
    with open(STD/fn, newline="") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        if r.get("dataset") in _UNMAP:
            r["dataset"] = _UNMAP[r["dataset"]]
    return rows
bw_std   = load_std("body_weight_long.csv")
food_std = load_std("food_intake_long.csv")
dev_std  = load_std("device_summary_long.csv")
subj_std = load_std("subjects.csv")

# ---------- prism helpers (independent extraction) ----------
def clean_title(t):
    if isinstance(t,dict): t=t.get("string","")
    return re.sub(r"[\r\n]+"," ",str(t)).strip()
def prism_sheet(prism, title):
    z=zipfile.ZipFile(prism); names=set(z.namelist())
    for n in sorted(names):
        m=re.match(r"data/sheets/([^/]+)/sheet\.json$",n)
        if not m: continue
        sh=json.loads(z.read(n)); tbl=sh.get("table") or {}
        if clean_title(sh.get("title",""))!=title: continue
        tuid=tbl.get("uid")
        if not tuid or f"data/tables/{tuid}/data.csv" not in names: continue
        dsets=[clean_title(json.loads(z.read(f"data/sets/{u}.json")).get("title")) for u in tbl.get("dataSets",[])]
        rows=[r.split(",") for r in z.read(f"data/tables/{tuid}/data.csv").decode("utf-8","replace").replace("\r\n","\n").split("\n")]
        return rows, dsets
    return None, None

def parse_date_any(s):
    s=str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S","%Y-%m-%d","%m/%d/%Y"):
        try: return datetime.strptime(s,fmt).date()
        except: pass
    return None

print("="*72); print("1) CYNTHIA body weight & food  <-  original updated .prism"); print("="*72)
def verify_cyn_timeline(prism_title, valfield, std_rows):
    rows,dsets=prism_sheet(CYN/"original/E4 D-TRF single-housed updated.prism", prism_title)
    mice=dsets                      # order == data columns 2..
    exp={}
    for r in rows:
        if len(r)<2+len(mice): continue
        d=parse_date_any(r[1])
        if d is None: continue
        for i,mo in enumerate(mice):
            v=fnum(r[2+i])
            if v is not None: exp[(mo,d)]=v
    got={(x["subject_id"],parse_date_any(x["date"])):fnum(x[valfield])
         for x in std_rows if x["dataset"]=="Cynthia"}
    mism=[k for k in exp if k not in got or abs(got[k]-exp[k])>TOL]
    extra=[k for k in got if k not in exp]
    check(f"{prism_title}: mouse count", set(mice)=={"1316","1338","1339","1340","1341","1342"}, str(mice))
    check(f"{prism_title}: cell count matches ({len(exp)} orig, {len(got)} std)", len(exp)==len(got))
    check(f"{prism_title}: all values match", not mism and not extra, f"mismatch={mism[:5]} extra={extra[:5]}")
verify_cyn_timeline("Weight timeline","body_weight_g",bw_std)
verify_cyn_timeline("Food timeline","food_g_per_mouse_per_day",food_std)

print("="*72); print("2) WINSTON body weight  <-  original xlsx (Copy of Body Weights)"); print("="*72)
wb=openpyxl.load_workbook(WIN/"original/[Temp File] Kravtiz Lab CohM2 1wk Mice Data Share.xlsx",data_only=True)
ws=wb["Copy of Body Weights"]
grid=[[c.value for c in row] for row in ws.iter_rows()]
hr=next(i for i,row in enumerate(grid) if str(row[0]).strip()=="Cage #")
datecols={j:(v.date() if isinstance(v,datetime) else parse_date_any(v)) for j,v in enumerate(grid[hr]) if j>=3 and v is not None}
exp={}
cur_cage=None
for row in grid[hr+1:]:
    if row[1] is None: continue
    mouse=str(row[1]).strip()
    for j,d in datecols.items():
        v=fnum(row[j])
        if v is not None: exp[(mouse,d)]=v
got={(x["subject_id"],parse_date_any(x["date"])):fnum(x["body_weight_g"]) for x in bw_std if x["dataset"]=="Winston"}
mism=[k for k in exp if k not in got or abs(got[k]-exp[k])>TOL]; extra=[k for k in got if k not in exp]
mice=set(m for m,_ in exp)
check("Winston BW: 22 mice", len(mice)==22, str(sorted(mice)))
check(f"Winston BW: cell count ({len(exp)} orig, {len(got)} std)", len(exp)==len(got))
check("Winston BW: all values match", not mism and not extra, f"mismatch={mism[:5]} extra={extra[:5]}")

print("="*72); print("3) WINSTON food  <-  BOTH original isocaloric .prism AND xlsx avg block"); print("="*72)
CAGE={"TRF_1":"E","TRF_2":"F","TRF_3":"I","TRF_4":"J","AL_1":"G","AL_2":"H","AL_3":"K"}
# (a) from prism
rows,dsets=prism_sheet(WIN/"original/Winston Li TRF Pre-Injury CohM2 Isocaloric Feeding.prism","Preinjury CohM2")
labels=["TRF_1","TRF_2","TRF_3","TRF_4","AL_1","AL_2","AL_3"]
exp_p={}
for r in rows:
    day=fnum(r[0])
    if day is None: continue
    for i,lab in enumerate(labels):
        v=fnum(r[1+i])
        if v is not None: exp_p[(CAGE[lab],int(day))]=v
got={(x["subject_id"],int(x["day_index"])):fnum(x["food_g_per_mouse_per_day"]) for x in food_std if x["dataset"]=="Winston"}
mism=[k for k in exp_p if k not in got or abs(got[k]-exp_p[k])>TOL]; extra=[k for k in got if k not in exp_p]
check(f"Winston food vs prism: cell count ({len(exp_p)} orig, {len(got)} std)", len(exp_p)==len(got))
check("Winston food vs prism: all values match (by day_index)", not mism and not extra, f"mismatch={mism[:5]} extra={extra[:5]}")
# (b) from xlsx avg block (cols J..P = 9..15; dates col I = 8) -> authoritative cage labels
wsf=wb["Copy of Food Hopper Tracking Ra"]
fg=[[c.value for c in row] for row in wsf.iter_rows()]
avg_cage_cols={9:"E",10:"F",11:"G",12:"H",13:"I",14:"J",15:"K"}   # 0-indexed J..P
exp_x={}
for row in fg[10:]:                      # data starts row 11 (idx 10)
    d=row[8].date() if isinstance(row[8],datetime) else parse_date_any(row[8])
    if d is None: continue
    for j,cg in avg_cage_cols.items():
        v=fnum(row[j])
        if v is not None: exp_x[(cg,d)]=v
gotx={(x["subject_id"],parse_date_any(x["date"])):fnum(x["food_g_per_mouse_per_day"]) for x in food_std if x["dataset"]=="Winston"}
mism=[k for k in exp_x if k not in gotx or abs(gotx[k]-exp_x[k])>TOL]; extra=[k for k in gotx if k not in exp_x]
check(f"Winston food vs xlsx: cell count ({len(exp_x)} orig, {len(gotx)} std)", len(exp_x)==len(gotx))
check("Winston food vs xlsx: all values match (by cage+date) -> cage mapping correct",
      not mism and not extra, f"mismatch={mism[:5]} extra={extra[:5]}")

print("="*72); print("4) DEVICE summary  <-  raw logs read directly from original zips"); print("="*72)
def zip_logs(zippath, unit_fn):
    z=zipfile.ZipFile(zippath); recs={}
    total=0
    for n in z.namelist():
        if not n.lower().endswith(".csv") or "__MACOSX" in n or os.path.basename(n).startswith("._"): continue
        txt=z.read(n).decode("utf-8","replace").replace("\r\n","\n").splitlines()
        if not txt: continue
        rdr=list(csv.DictReader(txt))
        if not rdr or "Datetime" not in rdr[0] or "Error" not in rdr[0]: continue
        unit=unit_fn(n, rdr)
        for row in rdr:
            d=parse_date_any(str(row["Datetime"]).split()[0])
            if d is None: continue
            total+=1
            k=(unit,d); a=recs.setdefault(k,{"n":0,"batt":[],"door":0,"err":0,"td":0,"ff":0})
            a["n"]+=1
            b=fnum(row["Battery_Voltage"]);  a["batt"].append(b) if b is not None else None
            dr=fnum(row["DoorOpen"]);         a["door"]+= (1 if dr==1 else 0)
            a["err"]+= (0 if str(row["Error"]).strip()=="OK" else 1)
            a["td"]+= (row["Task"]=="TimedDoor"); a["ff"]+= (row["Task"]=="FreeFeeding")
    return recs,total
def cyn_u(n,rdr): return "TUMBLY"+str(rdr[0]["Device_Number"]).zfill(3)
def win_u(n,rdr):
    m=re.search(r"Unit (\d+)",n); return f"Unit{m.group(1)}" if m else "NA"
cyn_recs,cyn_tot=zip_logs(CYN/"original/Tumbly csv logs.zip",cyn_u)
win_recs,win_tot=zip_logs(WIN/"original/TRF PreInjury Cohort M2 Tumbly SD Card File History.zip",win_u)
allrecs={("Cynthia",)+k:v for k,v in cyn_recs.items()}; allrecs.update({("Winston",)+k:v for k,v in win_recs.items()})
std_dev={(x["dataset"],x["unit"],parse_date_any(x["date"])):x for x in dev_std}
bad=[];
for k,a in allrecs.items():
    s=std_dev.get(k)
    if s is None: bad.append(("MISSING",k)); continue
    n_ok = int(float(s["n_samples"]))==a["n"]
    err_ok = abs(float(s["pct_error"]) - 100*a["err"]/a["n"])<0.01
    door_ok= abs(float(s["pct_door_open"]) - 100*a["door"]/a["n"])<0.01
    batt=[b for b in a["batt"] if b is not None]
    mb_ok = abs(float(s["mean_battery_v"]) - round(sum(batt)/len(batt),3))<0.01 if batt else True
    td_ok = int(float(s["n_timeddoor"]))==a["td"]; ff_ok=int(float(s["n_freefeeding"]))==a["ff"]
    if not all([n_ok,err_ok,door_ok,mb_ok,td_ok,ff_ok]): bad.append((k,dict(n_ok=n_ok,err_ok=err_ok,door_ok=door_ok,mb_ok=mb_ok,td_ok=td_ok,ff_ok=ff_ok)))
extra_keys=[k for k in std_dev if k not in allrecs]
check(f"Device: unit-day keys match ({len(allrecs)} orig, {len(std_dev)} std)", len(allrecs)==len(std_dev) and not extra_keys, f"extra_std={extra_keys[:5]}")
check("Device: every unit-day aggregate matches raw", not bad, str(bad[:3]))
sum_std=sum(int(float(x["n_samples"])) for x in dev_std)
check(f"Device: total sample completeness (raw {cyn_tot+win_tot} == std {sum_std})", cyn_tot+win_tot==sum_std)

print("="*72); print("5) day_index integrity (all tables)"); print("="*72)
di_bad=[]
for x in bw_std+food_std+dev_std:
    d=parse_date_any(x["date"]); ds=x["dataset"]
    if d and int(x["day_index"])!=(d-TRF_START[ds]).days: di_bad.append((ds,x.get("subject_id",x.get("unit")),x["date"]))
check("day_index == date - TRF_start everywhere", not di_bad, str(di_bad[:5]))

print("="*72); print("6) subjects roster"); print("="*72)
cyn_ids={x["subject_id"] for x in subj_std if x["dataset"]=="Cynthia"}
win_trf={x["subject_id"] for x in subj_std if x["dataset"]=="Winston" and x["diet"]=="TRF"}
win_ctl={x["subject_id"] for x in subj_std if x["dataset"]=="Winston" and x["diet"]=="control"}
check("Cynthia 6 mice", cyn_ids=={"1316","1338","1339","1340","1341","1342"})
check("Winston TRF n=12", len(win_trf)==12, str(sorted(win_trf)))
check("Winston control n=10", len(win_ctl)==10, str(sorted(win_ctl)))

print("\n"+"="*72)
print("RESULT:", "ALL CHECKS PASSED ✔" if not problems else f"{len(problems)} FAILURE(S)")
for p in problems: print("  -",p)
